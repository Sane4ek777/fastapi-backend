import sqlite3
import requests
import openpyxl
from bs4 import BeautifulSoup
from fastapi import HTTPException, Body
from config import DB_FILE, XLSX_FILE
from logger import logger
import xml.etree.ElementTree as ET  # 👈 Добавили ET
from config import DB_FILE, XML_FILE  # 👈 Добавили XML_FILE
import re
import unicodedata
from transliterate import translit

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

def generate_slug(name: str) -> str:
    """Генерирует slug из названия товара, заменяя кириллицу на латиницу."""
    # Преобразуем строку в нижний регистр
    name = name.lower()

    # Транслитерируем кириллицу в латиницу
    name = translit(name, 'ru', reversed=True)  # 'ru' - для русского языка, reversed=True для кириллицы в латиницу

    # Убираем все символы, кроме букв, цифр, дефисов и пробелов
    name = re.sub(r'[^a-zA-Z0-9\s-]', '', name)  # Оставляем только латиницу, цифры, пробелы и дефисы

    # Заменяем пробелы и нижние подчеркивания на дефисы
    name = re.sub(r'[\s_-]+', '-', name).strip()  # Заменяем пробелы и подчеркивания на дефисы

    return name


def generate_unique_slug(name, cursor):
    """Генерирует уникальный slug, добавляя суффиксы при необходимости"""
    base_slug = generate_slug(name)
    slug = base_slug
    counter = 1

    cursor.execute("SELECT COUNT(*) FROM products WHERE slug = ?", (slug,))
    while cursor.fetchone()[0] > 0:  # Пока slug уже есть, добавляем суффикс
        slug = f"{base_slug}-{counter}"
        counter += 1
        cursor.execute("SELECT COUNT(*) FROM products WHERE slug = ?", (slug,))

    return slug

def get_price_from_xlsx(row_idx: int):
    """Получает цену, РРЦ и артикул по номеру строки в XLSX"""
    logger.info(f"Getting price info from XLSX for row {row_idx}")
    wb = openpyxl.load_workbook(XLSX_FILE)
    sheet = wb.active

    row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]
    price_cell = row[9]
    price_rrc_cell = row[10]
    article = row[1].value if len(row) >= 2 else None

    price = price_cell.value if price_cell else None
    price_rrc = price_rrc_cell.value if price_rrc_cell else None

    logger.info(f"Found price {price}, price_rrc {price_rrc}, article {article} at row {row_idx}")
    return price, price_rrc, article

def scrape_product_data(url: str, row_idx: int):
    """Собирает данные о товаре, включая категории"""
    try:
        logger.info(f"Scraping product data from URL: {url}, row: {row_idx}")
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code != 200:
            logger.error(f"Error {response.status_code} while requesting {url}")
            raise HTTPException(status_code=400, detail=f"Ошибка {response.status_code} при запросе {url}")

        soup = BeautifulSoup(response.text, 'html.parser')

        # Название товара
        title_tag = soup.find('h2', class_='h2 hidden-tablet hidden-mobile')
        title = title_tag.get_text(strip=True) if title_tag else 'Без названия'

        # Цена и РРЦ по номеру строки
        price, price_rrc, article = get_price_from_xlsx(row_idx)

        # Изображения
        image_urls = []
        image_gallery = soup.find('div', class_='BigImagerGallery')
        if image_gallery:
            for img_tag in image_gallery.find_all('img'):
                src = img_tag.get('src')
                if src:
                    full_image_url = "https://www.diamir.su" + src if not src.startswith("http") else src
                    image_urls.append(full_image_url)

        # Описание
        description_tag = soup.find('div', class_='cover-text')
        description = description_tag.get_text(separator=' ', strip=True) if description_tag else ''

        # Категории
        categories = extract_categories(soup)
        category_id = save_categories_to_db(categories)

        # Характеристики
        attributes = scrape_attributes(soup, article)
        if not attributes:
            logger.warning(f"No attributes found for article {article}. Skipping product.")
            return None
        logger.info(f"Scraped data: {title}, Categories: {categories}, Price: {price}, Attributes: {attributes}")

        return {
            'name': title,
            'description': description,
            'price': price,
            'price_rrc': price_rrc,
            'images': image_urls,
            'category_id': category_id,
            'attributes': attributes
        }
    except HTTPException as http_error:
        logger.error(f"HTTP error occurred: {str(http_error)}")
        raise http_error
    except Exception as e:
        logger.error(f"Error occurred while scraping data from {url}: {str(e)}")
        return None

def update_product_names():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    # Выбираем только товары из файла Diamir
    cursor.execute("SELECT id, name FROM products WHERE is_diamir = 1")
    products = cursor.fetchall()

    for product_id, name in products:
        cursor.execute("SELECT attribute_name, attribute_value FROM product_attributes WHERE product_id = ?", (product_id,))
        attributes = cursor.fetchall()

        attribute_counts = {}
        attr_priority = {"зерн": None, "диаметр": None}

        for attr_name, attr_value in attributes:
            if attr_name == "Артикул для заказа":
                continue

            cleaned_value = attr_value.strip()

            # Запоминаем первую попавшуюся подходящую характеристику по приоритету
            lower_name = attr_name.lower()
            if "зерн" in lower_name and attr_priority["зерн"] is None:
                attr_priority["зерн"] = (attr_name, cleaned_value)
            elif "диаметр" in lower_name and attr_priority["диаметр"] is None:
                attr_priority["диаметр"] = (attr_name, cleaned_value)

            if attr_name not in attribute_counts:
                attribute_counts[attr_name] = set()
            attribute_counts[attr_name].add(cleaned_value)

        # Выбираем характеристику по приоритету
        significant_attr = None
        if attr_priority["зерн"]:
            significant_attr = attr_priority["зерн"]
        elif attr_priority["диаметр"]:
            significant_attr = attr_priority["диаметр"]
        else:
            max_unique_values = 0
            for attr_name, values in attribute_counts.items():
                if len(values) > max_unique_values:
                    max_unique_values = len(values)
                    significant_attr = (attr_name, next(iter(values)))

        # Обновляем имя, если нужно
        if significant_attr:
            attr_str = f"({significant_attr[0]}: {significant_attr[1]})"
            if attr_str not in name:
                new_name = f"{name} {attr_str}"
                cursor.execute("UPDATE products SET name = ? WHERE id = ?", (new_name, product_id))

    conn.commit()
    conn.close()
    
def scrape_all_products(urls_with_rows):
    """
    Обрабатывает все товары из списка URL и номеров строк
    urls_with_rows - список кортежей (url, row_idx)
    """
    for url, row_idx in urls_with_rows:
        product_data = scrape_product_data(url, row_idx)
        if product_data:
            try:
                insert_scraped_product(product_data)
                logger.info(f"Successfully added product from URL: {url}")
            except Exception as e:
                logger.error(f"Error inserting product data from URL {url}: {str(e)}")
        else:
            logger.warning(f"Skipping product due to scraping error: {url}")
    update_product_names()

def scrape_attributes(soup, article: str):
    """Собирает характеристики товара с веб-страницы"""
    logger.info(f"Scraping attributes for article: {article}")
    attributes = []

    article = str(article).strip().zfill(6)

    table_section = soup.find('section', class_='info__table table js-table')
    if not table_section:
        logger.warning("Attributes table not found!")
        return attributes

    coll_blocks = table_section.find_all('div', class_='table__coll coll')

    # Найдём блок с артикулами и определим, сколько их
    article_position = None
    article_values_count = 0
    for block in coll_blocks:
        name_tag = block.find('div', class_='coll__name')
        if name_tag and "Артикул для заказа" in name_tag.get_text(strip=True):
            container = block.find('div', class_='coll__container')
            value_divs = container.find_all('div', recursive=False)
            article_values_count = len(value_divs)

            # Если больше одного артикула — определим позицию
            if article_values_count > 1:
                for idx, val_div in enumerate(value_divs):
                    val_text = val_div.get_text(strip=True)
                    if val_text.zfill(6) == article:
                        article_position = idx
                        logger.info(f"Found article {article} at position {article_position}")
                        break
            break

    if article_values_count > 1 and article_position is None:
        logger.warning(f"Article {article} not found in attributes table.")
        return attributes

    # Сбор характеристик
    if article_values_count > 1:
        # Множественные артикулы — собираем по позиции
        for block in coll_blocks:
            name_tag = block.find('div', class_='coll__name')
            if not name_tag:
                continue
            name_text = name_tag.get_text(strip=True)
            container = block.find('div', class_='coll__container')
            value_divs = container.find_all('div', recursive=False)

            if len(value_divs) > article_position:
                value_text = value_divs[article_position].get_text(strip=True)
                attributes.append({'name': name_text, 'value': value_text})
                logger.info(f"Attribute parsed: {name_text} = {value_text}")

    else:
        # Один артикул — названия в первом блоке, значения во втором
        if len(coll_blocks) < 2:
            logger.warning("Expected two columns for single article, but found less.")
            return attributes

        # Названия характеристик
        name_block = coll_blocks[0]
        name_container = name_block.find('div', class_='coll__container')
        name_divs = name_container.find_all('div', recursive=False)

        # Значения характеристик
        value_block = coll_blocks[1]
        value_container = value_block.find('div', class_='coll__container')
        value_divs = value_container.find_all('div', recursive=False)

        # Сопоставим названия и значения
        for name_div, value_div in zip(name_divs, value_divs):
            name_text = name_div.get_text(strip=True)
            value_text = value_div.get_text(strip=True)
            attributes.append({'name': name_text, 'value': value_text})
            logger.info(f"Attribute parsed (single article): {name_text} = {value_text}")

    return attributes

def generate_unique_product_id():
    """Генерирует уникальный ID для товара без префикса."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    try:
        while True:
            # Извлекаем максимальный ID для товаров
            cursor.execute("SELECT MAX(id) FROM products")
            max_id = cursor.fetchone()[0]
            
            if max_id is None:
                new_id = 1  # Начинаем с 1 без префикса
            else:
                new_id = max_id + 1  # Следующий ID

            # Проверяем, существует ли товар с таким ID
            cursor.execute("SELECT 1 FROM products WHERE id = ?", (new_id,))
            if cursor.fetchone() is None:
                # Если товар с таким ID не существует, возвращаем ID
                conn.close()
                logger.info(f"Generated new product ID: {new_id}")
                return new_id

    except sqlite3.Error as e:
        logger.error(f"SQLite error: {str(e)}")
        conn.close()
        raise

    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        conn.close()
        raise

def insert_scraped_product(data):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    try:
        # 1. Извлечь артикул
        article = None
        for attr in data.get('attributes', []):
            if attr['name'].strip().lower() == "артикул для заказа":
                article = attr['value'].strip()
                break

        # 2. Проверка на существование товара
        if article:
            cursor.execute("""
                SELECT p.id FROM products p
                JOIN product_attributes a ON a.product_id = p.id
                WHERE a.attribute_name = 'Артикул для заказа' AND a.attribute_value = ?
            """, (article,))
            
            existing = cursor.fetchone()
            if existing:
                logger.warning(f"Товар с артикулом {article} уже существует (ID: {existing[0]}), пропускаем.")
                conn.close()
                return

        # 3. Продолжить добавление
        new_id = generate_unique_product_id()
        slug = generate_unique_slug(data['name'], cursor)

        cursor.execute(
            "INSERT INTO products (id, name, slug, price, price_rrc, description, image, category_id, available, is_diamir) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (new_id, data['name'], slug, data['price'], data['price_rrc'], data['description'], ",".join(data['images']), data['category_id'], True, True)
        )

        for attr in data.get('attributes', []):
            cursor.execute(
                "INSERT INTO product_attributes (product_id, attribute_name, attribute_value) VALUES (?, ?, ?)",
                (new_id, attr['name'], attr['value'])
            )

        conn.commit()
        logger.info(f"Товар '{data['name']}' (slug: {slug}) с ID {new_id} успешно добавлен.")
    except Exception as e:
        logger.error(f"Ошибка при добавлении товара: {str(e)}")
    finally:
        conn.close()

def normalize_name(name):
    """Приводит имя к нормализованному виду для сравнения"""
    return name.strip().lower()

def extract_categories(soup):
    """Извлекает иерархию категорий из хлебных крошек, исключая 'Главная', 'Каталог' и название товара"""
    breadcrumb_section = soup.find('section', class_='section breadcrumbs')
    if not breadcrumb_section:
        logger.warning("Breadcrumbs not found!")
        return []

    categories = []

    for link in breadcrumb_section.find_all('a', class_='breadcrumbs__item'):
        text = link.get_text(strip=True)
        href = link.get('href')

        # Пропускаем "Главная" и "Каталог"
        if text in ('Главная', 'Каталог'):
            continue

        # Берём только категории с href
        if href:
            categories.append(text)

    return categories

def save_categories_to_db(categories):
    """Сохраняет категории в БД с правильной иерархией parent_id"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    # Загружаем категории из XML
    tree = ET.parse(XML_FILE)
    root = tree.getroot()
    xml_categories = {
        normalize_name(category.text): int(category.get("id"))
        for category in root.findall(".//category") if category.text
    }

    parent_id = None  # ID родителя для текущей категории

    for category in categories:
        normalized = normalize_name(category)
        category_slug = generate_slug(category)

        # Проверка в XML
        if normalized in xml_categories:
            category_id = xml_categories[normalized]
        else:
            # Проверка в БД по имени
            cursor.execute("SELECT id FROM categories WHERE name = ?", (category,))
            result = cursor.fetchone()

            if result:
                category_id = result[0]
            else:
                # Проверка уникальности slug
                cursor.execute("SELECT COUNT(*) FROM categories WHERE slug = ?", (category_slug,))
                if cursor.fetchone()[0] > 0:
                    base_slug = category_slug
                    counter = 1
                    while True:
                        new_slug = f"{base_slug}-{counter}"
                        cursor.execute("SELECT COUNT(*) FROM categories WHERE slug = ?", (new_slug,))
                        if cursor.fetchone()[0] == 0:
                            category_slug = new_slug
                            break
                        counter += 1

                # Генерация ID (можно просто оставить автоинкремент, но на всякий случай сохраняем подход)
                cursor.execute("SELECT MAX(id) FROM categories")
                max_id = cursor.fetchone()[0] or 0
                new_category_id = max_id + 1

                try:
                    cursor.execute(
                        "INSERT INTO categories (id, name, slug, parent_id) VALUES (?, ?, ?, ?)",
                        (new_category_id, category, category_slug, parent_id)
                    )
                    conn.commit()
                    category_id = new_category_id
                    logger.info(f"Создана категория '{category}' с id={category_id} и parent_id={parent_id}")
                except sqlite3.IntegrityError:
                    logger.warning(f"ID {new_category_id} уже занят, вставляем без ID")
                    cursor.execute(
                        "INSERT INTO categories (name, slug, parent_id) VALUES (?, ?, ?)",
                        (category, category_slug, parent_id)
                    )
                    conn.commit()
                    category_id = cursor.lastrowid
                    logger.info(f"Создана категория '{category}' (autoid) с parent_id={parent_id}")

        parent_id = category_id  # Назначаем следующей категории этого как родителя

    conn.close()
    return parent_id  # Возвращаем ID последней (глубокой) категории

def is_url_valid(url: str) -> bool:
    """
    Проверяет, доступна ли страница по URL (код ответа не 404).
    """
    try:
        response = requests.head(url, allow_redirects=True, timeout=5)
        if response.status_code == 404:
            logger.warning(f"URL возвращает 404: {url}")
            return False
        logger.info(f"URL доступен: {url} (Status: {response.status_code})")
        return True
    except requests.RequestException as e:
        logger.error(f"Ошибка при проверке URL {url}: {str(e)}")
        return False

def parse_diamir_xlsx():
    """
    Извлекает ссылки на товары из XLSX-файла, проверяет их доступность и вызывает парсинг,
    передавая номер строки, в которой найдена гиперссылка
    """
    logger.info("Parsing Diamir XLSX file with hyperlink extraction and inline scraping...")
    wb = openpyxl.load_workbook(XLSX_FILE)
    sheet = wb.active
    found_any_links = False  # Флаг, который отслеживает, были ли найдены ссылки

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        for cell in row:
            if cell.value == "Ссылка на сайт" and cell.hyperlink:
                url = cell.hyperlink.target
                logger.info(f"Found hyperlink: {url} at row {row_idx}")
                found_any_links = True

                # Проверяем доступность URL перед парсингом
                if is_url_valid(url):
                    scrape_all_products([(url, row_idx)])
                else:
                    logger.warning(f"Пропуск недоступной ссылки: {url}")

    if not found_any_links:
        logger.warning("No links found in XLSX file.")

    logger.info("Finished parsing and scraping XLSX file.")

    